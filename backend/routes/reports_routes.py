"""Reports routes for analytics and data queries"""
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required
from sqlalchemy import and_, or_, func
from db import db
from models import Passenger
from utils.pagination import Paginator
from datetime import datetime

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/not-traveled', methods=['GET'])
@jwt_required()
def get_not_traveled_passengers():
    """
    Get passengers who booked a package but did NOT travel
    
    Query params:
    - package_name: Package name to filter by (REQUIRED)
    - from_date: Start date (optional, ISO format)
    - to_date: End date (optional, ISO format)
    - page: Page number (default 1)
    - per_page: Results per page (default 50)
    - search: Search by name/email (optional)
    """
    try:
        # Get filter parameters
        package_name = request.args.get('package_name', '').strip()
        from_date = request.args.get('from_date', '').strip()
        to_date = request.args.get('to_date', '').strip()
        search = request.args.get('search', '').strip()
        
        # Validate package_name is provided
        if not package_name:
            return jsonify({'error': 'package_name parameter is required'}), 400
        
        # Get pagination params
        page, per_page = Paginator.get_pagination_params()
        
        # Build base query
        query = Passenger.query.filter(Passenger.deleted_at.is_(None))
        
        # Filter by package name
        query = query.filter(
            Passenger.package_name.ilike(f'%{package_name}%')
        )
        
        # Filter by travel status - passengers who did NOT travel
        # Condition: travel_date IS NULL OR status is not 'completed'/'delivered'
        query = query.filter(
            or_(
                Passenger.journey_date.is_(None),
                ~func.lower(Passenger.status).ilike('%delivered%'),
                ~func.lower(Passenger.status).ilike('%completed%')
            )
        )
        
        # Apply date range filter (booking date)
        if from_date:
            try:
                from_datetime = datetime.fromisoformat(from_date)
                query = query.filter(Passenger.booking_date >= from_datetime)
            except:
                pass
        
        if to_date:
            try:
                to_datetime = datetime.fromisoformat(to_date)
                query = query.filter(Passenger.booking_date <= to_datetime)
            except:
                pass
        
        # Apply search filter
        if search:
            query = query.filter(
                or_(
                    Passenger.master_passenger_name.ilike(f'%{search}%'),
                    Passenger.email_id.ilike(f'%{search}%'),
                    Passenger.contact_number.ilike(f'%{search}%')
                )
            )
        
        # Sort by booking date descending
        query = query.order_by(Passenger.booking_date.desc())
        
        # Paginate
        result = Paginator.paginate_query(query, page, per_page)
        
        return jsonify(result), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/not-traveled/export', methods=['GET'])
@jwt_required()
def export_not_traveled_passengers():
    """
    Export not-traveled passengers to Excel
    
    Query params: Same as /not-traveled endpoint
    """
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        
        # Get filter parameters
        package_name = request.args.get('package_name', '').strip()
        from_date = request.args.get('from_date', '').strip()
        to_date = request.args.get('to_date', '').strip()
        search = request.args.get('search', '').strip()
        
        if not package_name:
            return jsonify({'error': 'package_name parameter is required'}), 400
        
        # Build query (same as not-traveled endpoint, but without pagination)
        query = Passenger.query.filter(Passenger.deleted_at.is_(None))
        
        query = query.filter(
            Passenger.package_name.ilike(f'%{package_name}%')
        )
        
        query = query.filter(
            or_(
                Passenger.journey_date.is_(None),
                ~func.lower(Passenger.status).ilike('%delivered%'),
                ~func.lower(Passenger.status).ilike('%completed%')
            )
        )
        
        if from_date:
            try:
                from_datetime = datetime.fromisoformat(from_date)
                query = query.filter(Passenger.booking_date >= from_datetime)
            except:
                pass
        
        if to_date:
            try:
                to_datetime = datetime.fromisoformat(to_date)
                query = query.filter(Passenger.booking_date <= to_datetime)
            except:
                pass
        
        if search:
            query = query.filter(
                or_(
                    Passenger.master_passenger_name.ilike(f'%{search}%'),
                    Passenger.email_id.ilike(f'%{search}%'),
                    Passenger.contact_number.ilike(f'%{search}%')
                )
            )
        
        passengers = query.order_by(Passenger.booking_date.desc()).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Not Traveled"
        
        # Define headers
        headers = [
            'Passenger Name',
            'Mobile',
            'Email',
            'Package Name',
            'Booking Date',
            'Expected Travel Date',
            'Status'
        ]
        
        # Add header row with formatting
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Set column widths
        column_widths = [25, 15, 25, 20, 15, 20, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Add data rows
        for row_num, passenger in enumerate(passengers, 2):
            ws.cell(row=row_num, column=1).value = passenger.master_passenger_name or ''
            ws.cell(row=row_num, column=2).value = passenger.contact_number or ''
            ws.cell(row=row_num, column=3).value = passenger.email_id or ''
            ws.cell(row=row_num, column=4).value = passenger.package_name or ''
            
            # Format booking date
            booking_date_cell = ws.cell(row=row_num, column=5)
            if passenger.booking_date:
                booking_date_cell.value = passenger.booking_date.strftime('%d/%m/%Y')
            else:
                booking_date_cell.value = ''
            
            # Format travel date
            travel_date_cell = ws.cell(row=row_num, column=6)
            if passenger.journey_date:
                travel_date_cell.value = passenger.journey_date.strftime('%d/%m/%Y')
            else:
                travel_date_cell.value = 'Not Traveled'
            
            ws.cell(row=row_num, column=7).value = passenger.status or ''
            
            # Center align status column
            ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="center")
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Convert to bytes
        bytes_output = BytesIO()
        wb.save(bytes_output)
        bytes_output.seek(0)
        
        filename = f"not_traveled_{package_name}_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            bytes_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/packages', methods=['GET'])
@jwt_required()
def get_all_packages():
    """Get list of all unique packages for dropdown"""
    try:
        # Get distinct package names
        packages = db.session.query(
            Passenger.package_name
        ).filter(
            Passenger.deleted_at.is_(None),
            Passenger.package_name.isnot(None)
        ).distinct().order_by(Passenger.package_name).all()
        
        package_list = [p[0] for p in packages if p[0]]
        
        return jsonify({
            'packages': package_list,
            'total': len(package_list)
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/merged-passengers', methods=['GET'])
@jwt_required()
def get_merged_passengers():
    """
    Get duplicate passengers grouped by name, mobile, or email
    Returns consolidated travel history for each duplicate group
    """
    try:
        search = request.args.get('search', '').strip().lower()
        
        # Find duplicates by name
        name_duplicates = db.session.query(
            func.lower(Passenger.master_passenger_name).label('name_key'),
            func.count(Passenger.id).label('count')
        ).filter(
            Passenger.deleted_at.is_(None),
            Passenger.master_passenger_name.isnot(None)
        ).group_by('name_key').having(func.count(Passenger.id) > 1).all()
        
        # Find duplicates by mobile
        mobile_duplicates = db.session.query(
            Passenger.contact_number,
            func.count(Passenger.id).label('count')
        ).filter(
            Passenger.deleted_at.is_(None),
            Passenger.contact_number.isnot(None)
        ).group_by(Passenger.contact_number).having(func.count(Passenger.id) > 1).all()
        
        # Find duplicates by email
        email_duplicates = db.session.query(
            func.lower(Passenger.email_id).label('email_key'),
            func.count(Passenger.id).label('count')
        ).filter(
            Passenger.deleted_at.is_(None),
            Passenger.email_id.isnot(None)
        ).group_by('email_key').having(func.count(Passenger.id) > 1).all()
        
        # Collect all duplicate keys
        duplicate_names = {d[0] for d in name_duplicates}
        duplicate_mobiles = {d[0] for d in mobile_duplicates}
        duplicate_emails = {d[0] for d in email_duplicates}
        
        # Find all passengers matching any duplicate criteria
        passengers = Passenger.query.filter(
            Passenger.deleted_at.is_(None),
            or_(
                func.lower(Passenger.master_passenger_name).in_(duplicate_names),
                Passenger.contact_number.in_(duplicate_mobiles),
                func.lower(Passenger.email_id).in_(duplicate_emails)
            )
        ).all()
        
        # Group passengers by duplicate criteria
        groups = {}
        
        for passenger in passengers:
            # Create a unique key for this group
            group_key = None
            
            # Check if matches any duplicate name
            if passenger.master_passenger_name:
                name_key = passenger.master_passenger_name.lower()
                if name_key in duplicate_names:
                    group_key = f"name_{name_key}"
            
            # Check if matches any duplicate mobile
            if not group_key and passenger.contact_number and passenger.contact_number in duplicate_mobiles:
                group_key = f"mobile_{passenger.contact_number}"
            
            # Check if matches any duplicate email
            if not group_key and passenger.email_id:
                email_key = passenger.email_id.lower()
                if email_key in duplicate_emails:
                    group_key = f"email_{email_key}"
            
            if group_key:
                if group_key not in groups:
                    groups[group_key] = []
                groups[group_key].append(passenger)
        
        # Build response
        merged_list = []
        
        for group_key, group_passengers in groups.items():
            if len(group_passengers) < 2:
                continue
            
            # Get master passenger (first one)
            master = group_passengers[0]
            
            # Collect all unique mobiles and emails
            mobiles = list(set(p.contact_number for p in group_passengers if p.contact_number))
            emails = list(set(p.email_id for p in group_passengers if p.email_id))
            
            # Collect all trips
            all_trips = []
            for p in group_passengers:
                all_trips.append({
                    'passenger_id': str(p.id),
                    'passenger_name': p.master_passenger_name,
                    'package': p.package_name or '',
                    'travel_date': p.journey_date.isoformat() if p.journey_date else None,
                    'booking_date': p.booking_date.isoformat() if p.booking_date else None,
                    'status': p.status or '',
                    'mobile': p.contact_number or '',
                    'email': p.email_id or ''
                })
            
            # Apply search filter
            if search:
                if (search not in master.master_passenger_name.lower() and
                    search not in str(mobiles).lower() and
                    search not in str(emails).lower()):
                    continue
            
            merged_list.append({
                'group_key': group_key,
                'master_name': master.master_passenger_name,
                'master_id': str(master.id),
                'mobiles': mobiles,
                'emails': emails,
                'duplicate_count': len(group_passengers),
                'total_trips': len(all_trips),
                'trips': sorted(all_trips, key=lambda x: x['booking_date'] or '', reverse=True)
            })
        
        # Sort by duplicate count
        merged_list.sort(key=lambda x: x['duplicate_count'], reverse=True)
        
        return jsonify({
            'data': merged_list,
            'total': len(merged_list)
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/merged-passengers/merge', methods=['POST'])
@jwt_required()
def merge_passenger_records():
    """
    Permanently merge duplicate passenger records
    Keeps master record and reassigns all trips to master
    """
    try:
        data = request.get_json()
        
        if not data or not data.get('group_key') or not data.get('master_id'):
            return jsonify({'error': 'group_key and master_id are required'}), 400
        
        group_key = data['group_key']
        master_id = data['master_id']
        
        # Start transaction
        try:
            # Get all passengers in this group
            search = request.args.get('search', '').strip().lower()
            
            # Find duplicates by name
            name_duplicates = db.session.query(
                func.lower(Passenger.master_passenger_name).label('name_key'),
                func.count(Passenger.id).label('count')
            ).filter(
                Passenger.deleted_at.is_(None),
                Passenger.master_passenger_name.isnot(None)
            ).group_by('name_key').having(func.count(Passenger.id) > 1).all()
            
            # Find duplicates by mobile
            mobile_duplicates = db.session.query(
                Passenger.contact_number,
                func.count(Passenger.id).label('count')
            ).filter(
                Passenger.deleted_at.is_(None),
                Passenger.contact_number.isnot(None)
            ).group_by(Passenger.contact_number).having(func.count(Passenger.id) > 1).all()
            
            # Find duplicates by email
            email_duplicates = db.session.query(
                func.lower(Passenger.email_id).label('email_key'),
                func.count(Passenger.id).label('count')
            ).filter(
                Passenger.deleted_at.is_(None),
                Passenger.email_id.isnot(None)
            ).group_by('email_key').having(func.count(Passenger.id) > 1).all()
            
            # Collect all duplicate keys
            duplicate_names = {d[0] for d in name_duplicates}
            duplicate_mobiles = {d[0] for d in mobile_duplicates}
            duplicate_emails = {d[0] for d in email_duplicates}
            
            # Find all passengers matching any duplicate criteria
            passengers = Passenger.query.filter(
                Passenger.deleted_at.is_(None),
                or_(
                    func.lower(Passenger.master_passenger_name).in_(duplicate_names),
                    Passenger.contact_number.in_(duplicate_mobiles),
                    func.lower(Passenger.email_id).in_(duplicate_emails)
                )
            ).all()
            
            # Find passengers in this specific group
            group_passengers = []
            
            for passenger in passengers:
                # Create a unique key for this group
                check_key = None
                
                # Check if matches any duplicate name
                if passenger.master_passenger_name:
                    name_key = passenger.master_passenger_name.lower()
                    if name_key in duplicate_names:
                        check_key = f"name_{name_key}"
                
                # Check if matches any duplicate mobile
                if not check_key and passenger.contact_number and passenger.contact_number in duplicate_mobiles:
                    check_key = f"mobile_{passenger.contact_number}"
                
                # Check if matches any duplicate email
                if not check_key and passenger.email_id:
                    email_key = passenger.email_id.lower()
                    if email_key in duplicate_emails:
                        check_key = f"email_{email_key}"
                
                if check_key == group_key:
                    group_passengers.append(passenger)
            
            if len(group_passengers) < 2:
                return jsonify({'error': 'Group must have at least 2 passengers'}), 400
            
            # Get master passenger
            master = Passenger.query.filter_by(id=master_id).first()
            if not master or master not in group_passengers:
                return jsonify({'error': 'Invalid master passenger'}), 400
            
            # Soft delete other passengers
            for passenger in group_passengers:
                if passenger.id != master_id:
                    passenger.deleted_at = datetime.utcnow()
            
            db.session.commit()
            
            return jsonify({
                'message': f'Successfully merged {len(group_passengers) - 1} duplicate records',
                'master_id': str(master_id)
            }), 200
        
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Merge failed: {str(e)}'}), 500
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
