from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from io import StringIO, BytesIO
import csv
import json
from db import db
from models import Passenger, User
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func

export_bp = Blueprint('export', __name__)

@export_bp.route('/csv', methods=['GET'])
@jwt_required()
def export_csv():
    """Export passengers as CSV"""
    user_id = get_jwt_identity()
    user = User.query.filter_by(id=user_id).first()
    
    if not user or user.role == 'viewer':
        return jsonify({'error': 'Permission denied. Viewers cannot export data.'}), 403

    try:
        # Get filter params (same as search)
        name = request.args.get('name', '').strip()
        city = request.args.get('city', '').strip()
        state = request.args.get('state', '').strip()
        status = request.args.get('status', '').strip()
        international = request.args.get('international', '').strip().lower()
        
        # Build query
        query = Passenger.query.filter(Passenger.deleted_at.is_(None))
        
        if name:
            query = query.filter(Passenger.master_passenger_name.ilike(f'%{name}%'))
        if city:
            query = query.filter(Passenger.city.ilike(f'%{city}%'))
        if state:
            query = query.filter(Passenger.state.ilike(f'%{state}%'))
        if status and status.lower() != 'all':
            query = query.filter(Passenger.status.ilike(f'%{status}%'))
        if international and international != 'all':
            intl_value = international in ['yes', 'true', '1']
            query = query.filter(Passenger.international_client == intl_value)
        
        passengers = query.all()
        
        # Create CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        headers = [
            'ID', 'User ID', 'Transaction ID', 'Passenger Name', 'Age', 'DOB',
            'Anniversary Date', 'Gender', 'Email', 'Contact', 'No. of Passengers',
            'Booking Date', 'Journey Date', 'Boarding Point', 'City', 'State',
            'Package Code', 'Package Name', 'Package Class', 'Status',
            'Nominee Name', 'Nominee Relation', 'Nominee Contact', 'International'
        ]
        writer.writerow(headers)
        
        # Write data
        for p in passengers:
            writer.writerow([
                str(p.id),
                p.user_id,
                p.transcation_id,
                p.master_passenger_name,
                p.age or '',
                p.dob.strftime('%d/%m/%Y') if p.dob else '',
                p.anniversary_date.strftime('%d/%m/%Y') if p.anniversary_date else '',
                p.gender or '',
                p.email_id or '',
                p.contact_number or '',
                p.no_of_passenger or '',
                p.booking_date.strftime('%d/%m/%Y') if p.booking_date else '',
                p.journey_date.strftime('%d/%m/%Y') if p.journey_date else '',
                p.boarding_point or '',
                p.city,
                p.state,
                p.package_code or '',
                p.package_name or '',
                p.package_class or '',
                p.status,
                p.nominee_name or '',
                p.nominee_relation or '',
                p.nominee_contact or '',
                'Yes' if p.international_client else 'No'
            ])
        
        # Convert to bytes
        output.seek(0)
        bytes_output = BytesIO()
        bytes_output.write(output.getvalue().encode('utf-8'))
        bytes_output.seek(0)
        
        filename = f"passengers_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return send_file(
            bytes_output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/json', methods=['GET'])
@jwt_required()
def export_json():
    """Export passengers as JSON"""
    user_id = get_jwt_identity()
    user = User.query.filter_by(id=user_id).first()
    
    if not user or user.role == 'viewer':
        return jsonify({'error': 'Permission denied. Viewers cannot export data.'}), 403

    try:
        # Get filter params
        name = request.args.get('name', '').strip()
        city = request.args.get('city', '').strip()
        state = request.args.get('state', '').strip()
        status = request.args.get('status', '').strip()
        international = request.args.get('international', '').strip().lower()
        
        # Build query
        query = Passenger.query.filter(Passenger.deleted_at.is_(None))
        
        if name:
            query = query.filter(Passenger.master_passenger_name.ilike(f'%{name}%'))
        if city:
            query = query.filter(Passenger.city.ilike(f'%{city}%'))
        if state:
            query = query.filter(Passenger.state.ilike(f'%{state}%'))
        if status and status.lower() != 'all':
            query = query.filter(Passenger.status.ilike(f'%{status}%'))
        if international and international != 'all':
            intl_value = international in ['yes', 'true', '1']
            query = query.filter(Passenger.international_client == intl_value)
        
        passengers = query.all()
        
        # Create JSON
        data = {
            'export_date': datetime.utcnow().isoformat(),
            'total_records': len(passengers),
            'passengers': [p.to_dict() for p in passengers]
        }
        
        bytes_output = BytesIO()
        bytes_output.write(json.dumps(data, indent=2).encode('utf-8'))
        bytes_output.seek(0)
        
        filename = f"passengers_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
        
        return send_file(
            bytes_output,
            mimetype='application/json',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/excel', methods=['GET'])
@jwt_required()
def export_excel():
    """Export passengers as Excel (.xlsx)"""
    user_id = get_jwt_identity()
    user = User.query.filter_by(id=user_id).first()
    
    if not user or user.role == 'viewer':
        return jsonify({'error': 'Permission denied. Viewers cannot export data.'}), 403

    try:
        # Get filter params (same as search)
        name = request.args.get('name', '').strip()
        city = request.args.get('city', '').strip()
        state = request.args.get('state', '').strip()
        status = request.args.get('status', '').strip()
        international = request.args.get('international', '').strip().lower()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        
        # Build query
        query = Passenger.query.filter(Passenger.deleted_at.is_(None))
        
        if name:
            query = query.filter(Passenger.master_passenger_name.ilike(f'%{name}%'))
        if city:
            query = query.filter(Passenger.city.ilike(f'%{city}%'))
        if state:
            query = query.filter(Passenger.state.ilike(f'%{state}%'))
        if status and status.lower() != 'all':
            query = query.filter(Passenger.status.ilike(f'%{status}%'))
        if international and international != 'all':
            intl_value = international in ['yes', 'true', '1']
            query = query.filter(Passenger.international_client == intl_value)
        
        # Apply date range filters if provided
        if date_from:
            try:
                from_date = datetime.fromisoformat(date_from)
                query = query.filter(Passenger.journey_date >= from_date)
            except:
                pass
        
        if date_to:
            try:
                to_date = datetime.fromisoformat(date_to)
                query = query.filter(Passenger.journey_date <= to_date)
            except:
                pass
        
        # Order by booking date descending
        passengers = query.order_by(Passenger.booking_date.desc()).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Passengers"
        
        # Define headers
        headers = [
            'Passenger Name',
            'Mobile',
            'Email',
            'City',
            'State',
            'Package',
            'Travel Date',
            'Status',
            'Created Date'
        ]
        
        # Add header row with formatting
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Set column widths
        column_widths = [25, 15, 25, 15, 15, 20, 15, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Add data rows
        for row_num, passenger in enumerate(passengers, 2):
            ws.cell(row=row_num, column=1).value = passenger.master_passenger_name or ''
            ws.cell(row=row_num, column=2).value = passenger.contact_number or ''
            ws.cell(row=row_num, column=3).value = passenger.email_id or ''
            ws.cell(row=row_num, column=4).value = passenger.city or ''
            ws.cell(row=row_num, column=5).value = passenger.state or ''
            ws.cell(row=row_num, column=6).value = passenger.package_name or ''
            
            # Format travel date
            travel_date_cell = ws.cell(row=row_num, column=7)
            if passenger.journey_date:
                travel_date_cell.value = passenger.journey_date.strftime('%d/%m/%Y')
            else:
                travel_date_cell.value = ''
            
            ws.cell(row=row_num, column=8).value = passenger.status or ''
            
            # Format created date
            created_date_cell = ws.cell(row=row_num, column=9)
            if passenger.created_at:
                created_date_cell.value = passenger.created_at.strftime('%d/%m/%Y')
            else:
                created_date_cell.value = ''
            
            # Center align status column
            ws.cell(row=row_num, column=8).alignment = Alignment(horizontal="center")
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Convert to bytes
        bytes_output = BytesIO()
        wb.save(bytes_output)
        bytes_output.seek(0)
        
        filename = f"passenger_export_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            bytes_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
